from openpyxl import *
from pywinauto import *
from pywinauto.keyboard import send_keys


def read_excel_families():
    wb = load_workbook('productos.xlsx')
    print(wb.sheetnames)
    print(wb.active)
    wb.active = wb['Familias']
    ws = wb.active

    cell_range = ws['A2':'A5']
    print(cell_range)
    cod_family = []
    family_name = []
    for row in ws.iter_cols(min_col=1, max_col=1, max_row=5, min_row=2, values_only=True):
        for cell in row:
            cod_family.append(cell)
            print(cell)

    for row in ws.iter_cols(min_col=2, max_col=2, max_row=5, min_row=2, values_only=True):
        for cell in row:
            family_name.append(cell)
            print(cell)
    return cod_family, family_name


def init_stock_app():
    app = Application(backend="uia").start('Stock/Stock.exe')
    # .connect(title='Control de deposito',timeout=10)
    menu_maintenance = app.ControlDeDeposito.child_window(title="Mantenimiento",
                                                          control_type="MenuItem").wrapper_object()
    menu_maintenance.click_input()
    send_keys('{DOWN}{DOWN}{ENTER}')
    app.ControlDeDeposito.print_control_identifiers()

    return app


def set_family_data():
    app = init_stock_app()
    family_id = app.ControlDeDeposito.child_window(auto_id="4", control_type="Edit").wrapper_object()
    family_name = app.ControlDeDeposito.child_window(auto_id="5", control_type="Edit").wrapper_object()
    save_button = app.ControlDeDeposito.child_window(title="Guardar", auto_id="3",
                                                    control_type="Button").wrapper_object()

    index = 0
    cod_family, name_family = read_excel_families()
    while index < len(cod_family):
        family_id.type_keys(cod_family[index])
        family_name.type_keys(name_family[index], with_spaces=True)
        save_button.click_input()
        index += 1
        family_id.set_text('')
        family_name.set_text('')
    close_maintenance = app.ControlDeDeposito.child_window(title="Salir", auto_id="2",
                                                           control_type="Button").wrapper_object()
    close_maintenance.click_input()


# agregar articulos
def read_excel_products():
    wb = load_workbook('productos.xlsx')
    wb.active = wb['Productos']
    ws = wb.active
    print(wb.active)
    productos = []
    for row in ws.iter_cols(min_row=2, values_only=True):
        for cell in row:
            productos.append(cell)
            print(cell)

    print(productos, len(productos))
    return productos


def set_data_products():
    app = Application(backend="uia").connect(title='Control de deposito', path='Stock/Stock.exe')
    # Application().connect(title='Control de deposito', class_name='Stock.exe')
    maintenance_menu = app.ControlDeDeposito.child_window(title="Mantenimiento",control_type="MenuItem").wrapper_object()
    maintenance_menu.click_input()
    send_keys('{DOWN}{ENTER}')
    app.ControlDeDeposito.print_control_identifiers()
    id_number = app.ControlDeDeposito.child_window(auto_id="7", control_type="Edit").wrapper_object()
    product_name = app.ControlDeDeposito.child_window(auto_id="8", control_type="Edit").wrapper_object()
    product_family = app.ControlDeDeposito.child_window(auto_id="3", control_type="Edit").wrapper_object()
    search_button = app.ControlDeDeposito.child_window(title="Buscar", auto_id="2",control_type="Button").wrapper_object()
    product_details = app.ControlDeDeposito.child_window(auto_id="6", control_type="Edit").wrapper_object()
    product_price = app.ControlDeDeposito.child_window(auto_id="9", control_type="Edit").wrapper_object()
    save_button = app.ControlDeDeposito.child_window(title="Guardar", auto_id="10",control_type="Button").wrapper_object()
    productos = read_excel_products()
    index = 0
    while index <= 9:
        id_number.type_keys(productos[index])
        product_name.type_keys(productos[index + 10], with_spaces=True)
        product_details.type_keys(productos[index + 30], with_spaces=True)
        product_price.type_keys(productos[index + 40])
        product_family.type_keys(productos[index + 20], with_spaces=True)
        search_button.click_input()
        handle = findwindows.find_window(best_match='Buscar algo')
        search_app = Application(backend="win32").connect(handle=handle)
        search_app.BuscarAlgo.click_input()
        send_keys('%{F4}')
        save_button.click_input()
        index += 1
        id_number.set_text('')
        product_name.set_text('')
        product_family.set_text('')
        product_details.set_text('')
        product_price.set_text('')


set_family_data()
set_data_products()
